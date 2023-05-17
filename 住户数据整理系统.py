# 住户数据整理系统 20230328
from tkinter import *
from tkinter.ttk import *
from tkinter import filedialog
from tkinter import messagebox
import pandas as pd
from sqlalchemy import create_engine
import pyodbc
import os
import openpyxl
"""
全局通用函数
"""
# 定义数据库服务器连接
myserver = 'DRIVER={ODBC Driver 17 for SQL Server};SERVER=10.132.60.188;DATABASE=hhgzhdc;ENCRYPT=yes;TrustServerCertificate=yes;UID=sa;PWD=zpf13062990859'
conn = pyodbc.connect(myserver)
cursor = conn.cursor()
engine = create_engine(
    'mssql+pymssql://sa:zpf13062990859@10.132.60.188:1433/hhgzhdc?charset=utf8')
conn2 = engine.connect()
# 定义时间范围
yearitems = ("2022", "2023", "2024", "2025", "2026")
monthitems = ("01", "02", "03", "04", "05", "06",
              "07", "08", "09", "10", "11", "12")

# 定义存档子程序

def save_to_excel(df, wjm, sheet_name):
    with pd.ExcelWriter(wjm, mode='a') as writer: # 使用pandas库中的ExcelWriter方法，将数据保存到Excel文件中
        df.to_excel(writer, sheet_name=sheet_name, index=False) # 将数据写入Excel文件的指定工作表中，不包含行索引
  
# 自动隐藏滚动条

def scrollbar_autohide(bar, widget):
    def show():
        bar.lift(widget)

    def hide():
        bar.lower(widget)
    hide()
    widget.bind("<Enter>", lambda e: show())
    bar.bind("<Enter>", lambda e: show())
    widget.bind("<Leave>", lambda e: hide())
    bar.bind("<Leave>", lambda e: hide())


class WinGUI(Tk):
    def __init__(self):
        super().__init__()
        self.__win()
        self.tk_label_label1 = self.__tk_label_label1()
        self.tk_select_box_checkbox1 = self.__tk_select_box_checkbox1()
        self.tk_label_label2 = self.__tk_label_label2()
        self.tk_label_label4 = self.__tk_label_label4()
        self.tk_select_box_checkbox2 = self.__tk_select_box_checkbox2()
        self.tk_label_label3 = self.__tk_label_label3()
        self.tk_select_box_checkbox3 = self.__tk_select_box_checkbox3()
        self.tk_button_button1 = self.__tk_button_button1()
        self.tk_button_button2 = self.__tk_button_button2()
        self.tk_button_button3 = self.__tk_button_button3()

    def __win(self):
        self.title("住户数据整理系统")
        # 设置窗口大小、居中
        width = 1024
        height = 768
        screenwidth = self.winfo_screenwidth()
        screenheight = self.winfo_screenheight()
        geometry = '%dx%d+%d+%d' % (width, height,
                                    (screenwidth - width) / 2, (screenheight - height) / 2)
        self.geometry(geometry)
        self.resizable(width=False, height=False)

    def __tk_label_label1(self):
        label = Label(self, text="请选择", anchor="center")
        label.place(x=20, y=40, width=50, height=24)
        return label

    def __tk_select_box_checkbox1(self):
        cb = Combobox(self, state="readonly")
        cb['values'] = yearitems
        cb.place(x=80, y=40, width=60, height=24)
        return cb

    def __tk_label_label2(self):
        label = Label(self, text="年", anchor="center")
        label.place(x=140, y=40, width=20, height=24)
        return label

    def __tk_label_label4(self):
        label = Label(self, text="请选择乡镇", anchor="center")
        label.place(x=20, y=90, width=80, height=24)
        return label

    def __tk_select_box_checkbox2(self):
        cb = Combobox(self, state="readonly")
        cb['values'] = monthitems
        cb.place(x=165, y=40, width=40, height=24)
        return cb

    def __tk_label_label3(self):
        label = Label(self, text="月", anchor="center")
        label.place(x=210, y=40, width=30, height=24)
        return label

    def __tk_select_box_checkbox3(self):
        cb = Combobox(self, state="readonly")
        cb['values'] = ("滨江镇", "分界镇")
        cb.place(x=120, y=90, width=80, height=24)
        return cb

    def __tk_button_button1(self):
        btn = Button(self, text="生成未编码表格")
        btn.place(x=30, y=170, width=130, height=24)
        return btn

    def __tk_button_button2(self):
        btn = Button(self, text="导入已编码表格")
        btn.place(x=30, y=220, width=130, height=24)
        return btn

    def __tk_button_button3(self):
        btn = Button(self, text="生成电子台帐")
        btn.place(x=30, y=270, width=130, height=24)
        return btn


class Win(WinGUI):
    def __init__(self):
        super().__init__()
        self.__event_bind()
# 生成未编码

    def scwbm(self, evt):
        # 更新帐目名称到备注note字段
        sql1 = "UPDATE  调查点台账合并 SET note=note+type_name COLLATE database_default,ybz='1' where year=" + "\'" + \
            win.tk_select_box_checkbox1.get() + "\'" + " and month = " + "\'" + \
            win.tk_select_box_checkbox2.get() + "\'" + " and ybz<>'1'"
        cursor.execute(sql1)
        # 读取未编码的数据
        sql2 = "SELECT   调查点户名单.户代码, 调查点户名单.户主姓名, 调查点台账合并.type_name, 调查点台账合并.amount AS 数量, 调查点台账合并.[date] as 日期," + \
            "调查点台账合并.money AS 金额, 调查点台账合并.note AS 备注, 调查点台账合并.type AS 收支, 调查点台账合并.id," + "调查点台账合并.code " + \
            "FROM 调查点台账合并 INNER JOIN  调查点户名单 ON 调查点台账合并.hudm = 调查点户名单.户代码 " + \
            "where 调查点台账合并.year = " + "\'" + win.tk_select_box_checkbox1.get() + "\'" + " and 调查点台账合并.month = " + "\'" + \
            win.tk_select_box_checkbox2.get() + "\'" + \
            " AND 调查点台账合并.code IS NULL ORDER BY 调查点台账合并.type_name"
        print(sql2)

        df = pd.read_sql(sql2, conn)
        wjm = "/home/zpf/" + win.tk_select_box_checkbox2.get()+"月未编码"+".xlsx"
        df.to_excel(wjm, index=False, sheet_name='Sheet1')
        excel_file = pd.ExcelFile(wjm)
        df = excel_file.parse()
        df.iloc[:, 0] = df.iloc[:, 0].astype(str)

        # Save changes and close the file
        writer = pd.ExcelWriter(wjm)
        df.to_excel(writer, index=False)
        writer.save()
        writer.close()


# 导入已编码数据

    def drybm(self, evt):
        # 打开需要导入的文件
        wbmbiaoge = filedialog.askopenfilename(
            initialdir="/home/zpf/", title="选择文件", filetypes=(("office 2003格式电子表格", "*.xlsx"), ("all files", "*.*")))
        df = pd.read_excel(wbmbiaoge, sheet_name='Sheet1')
        # Write to SQL Server
        df.to_sql('滨江补充编码1', engine, if_exists='replace')
        sql3 = "UPDATE 调查点台账合并 SET code = 滨江补充编码1.code,ybm='1' FROM 调查点台账合并 INNER JOIN 滨江补充编码1 ON 调查点台账合并.id = 滨江补充编码1.id"
        sql4 = "update 调查点台账合并 set [note] = [type_name]+[note] COLLATE database_default,ybz='1' " + \
            "where year = " + "\'" + win.tk_select_box_checkbox1.get()+"\'" + " and month= " + "\'"+win.tk_select_box_checkbox2.get()+"\'" + \
            " AND " + "(ybz<>'1')" + " and " + \
            "( 调查点台账合并.id  in (select distinct id from 滨江补充编码1))"
        sql5 = "UPDATE  调查点台账合并 SET type_name = 调查品种编码.帐目指标名称, unit_name = 调查品种编码.单位名称 FROM 调查点台账合并 INNER " +\
            "JOIN 调查品种编码 ON 调查点台账合并.code = 调查品种编码.帐目编码 " +\
            "where year = " + "\'" + win.tk_select_box_checkbox1.get()+"\'" + " and month= " + "\'"+win.tk_select_box_checkbox2.get()+"\'" + \
            " AND (调查点台账合并.code IS not NULL)"
        cursor.execute(sql3)
        print(cursor.rowcount)
        cursor.execute(sql4)
        print(cursor.rowcount)
        cursor.execute(sql5)
        print(cursor.rowcount)
        conn.commit()

# 生成电子台帐

    def scdztz(self, evt):
        #定义乡镇码
        xzm = {
            '分界镇':'321283101',
            '滨江镇':'321283123'
        }.get(win.tk_select_box_checkbox3.get(),'')
        #定义总天数
        zts = {
            '01': 31,
            '02': 28,
            '03': 31,
            '04': 30,
            '05': 31,
            '06': 30,
            '07': 31,
            '08': 31,
            '09': 30,
            '10': 31,
            '11': 30,
            '12': 31
        }.get(win.tk_select_box_checkbox2.get(),'')
        wjm = "/home/zpf/" + win.tk_select_box_checkbox3.get()+ win.tk_select_box_checkbox2.get() + "电子台帐" + ".xlsx"        
        sj1 = "year = " + "\'" + win.tk_select_box_checkbox1.get()+"\'" + " and month= " + \
            "\'"+win.tk_select_box_checkbox2.get()+"\'"
        sj2 = "调查点台账合并_2.year = " + "\'" + win.tk_select_box_checkbox1.get()+"\'" + \
            " and 调查点台账合并_2.month= " + "\'"+win.tk_select_box_checkbox2.get()+"\'"    
        sql6 = '''SELECT   derivedtbl_1.hudm AS 户代码, 调查点户名单.户主姓名, shourubiao.收入, derivedtbl_1.支出, derivedtbl_2.记账笔数, 
                derivedtbl_3.漏记账天数
FROM      (SELECT   hudm, ROUND(SUM(money), 0) AS 支出
                 FROM      调查点台账合并 AS 调查点台账合并_1
                 WHERE ''' + sj1+'''AND (type = '2') AND (hudm like \''''+xzm+'''%%%%%\')
                 GROUP BY hudm) AS derivedtbl_1 INNER JOIN
                调查点户名单 ON derivedtbl_1.hudm = 调查点户名单.户代码 INNER JOIN
                    (SELECT   调查点台账合并_2.hudm AS 户代码, 调查点户名单_1.户主姓名, COUNT(调查点台账合并_2.hudm) 
                                     AS 记账笔数
                     FROM      调查点台账合并 AS 调查点台账合并_2 INNER JOIN
                                     调查点户名单 AS 调查点户名单_1 ON 调查点台账合并_2.hudm = 调查点户名单_1.户代码
                     WHERE ''' + sj2 + '''
                     GROUP BY 调查点台账合并_2.hudm, 调查点户名单_1.户主姓名) AS derivedtbl_2 ON 
                derivedtbl_1.hudm = derivedtbl_2.户代码 INNER JOIN
                    (SELECT   调查点户名单_2.户主姓名, '''+str(zts)+''' - COUNT(table_1.hudm) AS 漏记账天数
                     FROM      (SELECT DISTINCT hudm, date
                                      FROM      调查点台账合并 AS 调查点台账合并_3
                                      WHERE '''+sj1 + ''') AS table_1 INNER JOIN
                                     调查点户名单 AS 调查点户名单_2 ON table_1.hudm = 调查点户名单_2.户代码
                     GROUP BY table_1.hudm, 调查点户名单_2.户主姓名) AS derivedtbl_3 ON 
                调查点户名单.户主姓名 = derivedtbl_3.户主姓名 LEFT OUTER JOIN
                    (SELECT   hudm, SUM(money) AS 收入
                     FROM      调查点台账合并
                     WHERE  '''+sj1+''' AND (type = '1') 
                     GROUP BY hudm) AS shourubiao ON derivedtbl_1.hudm = shourubiao.hudm'''
        sql7='''SELECT   调查点台账合并.hudm, 调查点户名单.户主姓名, 调查点台账合并.code, 调查点台账合并.amount as 数量, 
                调查点台账合并.money as 金额,  调查点台账合并.date, 调查点台账合并.type, 调查点台账合并.id, 调查点台账合并.type_name, 
                调查点台账合并.unit_name
FROM      调查点台账合并 INNER JOIN
                调查点户名单 ON 调查点台账合并.hudm = 调查点户名单.户代码
WHERE '''+sj1+''' and code is not null AND (hudm like \''''+xzm+'''%%%%%\')
order by hudm,type,date
'''
        sql8='''SELECT   derivedtbl_1.hudm, 调查点户名单.户主姓名,derivedtbl_1.code, 调查品种编码.帐目指标名称, round(derivedtbl_1.总金额,0) as 总金额,
                derivedtbl_1.记账笔数
FROM      调查品种编码 INNER JOIN
                    (SELECT DISTINCT COUNT(code) AS 记账笔数, SUM(money) AS 总金额, code, hudm
                     FROM      调查点台账合并
                     WHERE   left(code,2)<>'31' and (hudm like \''''+xzm+'''%%%%%\') and '''+sj1 +''' 
                     GROUP BY hudm, code) AS derivedtbl_1 ON 调查品种编码.帐目编码 = derivedtbl_1.code INNER JOIN
                调查点户名单 ON derivedtbl_1.hudm = 调查点户名单.户代码
order by hudm,code
'''
        # 读取数据
        df1 = pd.read_sql(sql6, conn)
        df2 = pd.read_sql(sql7, conn)
        df3 = pd.read_sql(sql8, conn)
        # 保存到 Excel 文件
        if os.path.exists(wjm):
            os.remove(wjm)
        workbook = openpyxl.Workbook()
        if 'Sheet1' in workbook.sheetnames:
            sheet = workbook['Sheet1']
            workbook.remove(sheet)
        workbook.save(wjm)
        save_to_excel(df1, wjm, '汇总表')
        save_to_excel(df2, wjm, '分户详细账')
        save_to_excel(df3, wjm, '分户消费结构')
        if 'Sheet1' in workbook.sheetnames:
            sheet = workbook['Sheet1']
            workbook.remove(sheet)
        # 关闭数据库连接
        messagebox.showinfo('完成', '已经生成台帐')
        
        
    def __event_bind(self):
        self.tk_button_button1.bind('<Button>', self.scwbm)
        self.tk_button_button2.bind('<Button>', self.drybm)
        self.tk_button_button3.bind('<Button>', self.scdztz)


if __name__ == "__main__":
    win = Win()
    win.mainloop()
